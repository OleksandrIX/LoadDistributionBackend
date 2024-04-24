import os
import re
import math
from loguru import logger
from openpyxl import load_workbook
from modules.excel_parser.utils.dataframe_utils import *
from modules.excel_parser.utils.spreadsheet_utils import *
from modules.excel_parser.utils.data_utils import *
from modules.excel_parser.exceptions.ParsingException import ParsingException


def get_data_from_header_block(header_block_worksheet: Worksheet) -> dict[str, str]:
    """Parsed header block in worksheet"""
    header_data = {
        "Рівень вищої освіти": None,
        "Спеціальність": None,
        "Спеціалізація": None,
        "Курс навчання": None,
        "Ступінь вищої освіти": None,
        "Навчальна група": None
    }
    for row in header_block_worksheet:
        cell_key = None
        for cell in row:
            cell_value = cell.value

            if cell_key and cell_value:
                header_data[cell_key] = cell_value
                cell_key = None

            if cell_value and isinstance(cell_value, str):
                cell_value = cell.value.strip()
                for key in header_data:
                    if cell_value.startswith(key):
                        cell_key = key
    return header_data


def get_data_from_table_block(table_dataframe: pd.DataFrame) -> list[dict]:
    """Parsed table block in worksheet"""
    table_data = []
    for _, row in table_dataframe.iterrows():
        education_component = {}

        for column, value in row.items():
            key = get_column_name_by_number(column)

            if key in ["department", "education_component_name", "education_component_code", "credits", "hours"]:
                education_component[key] = value
            else:
                semester_number = get_semester_number_by_column(column)

                if "semesters" not in education_component:
                    education_component["semesters"] = []
                if not semester_number - 1 < len(education_component["semesters"]):
                    education_component["semesters"].append({})

                semester = education_component["semesters"][semester_number - 1]
                semester["semester_number"] = semester_number

                if math.isnan(value):
                    value = 0

                if key == "total_amount_hours":
                    semester[key] = value
                elif key in ["amount_classroom_hours", "lecture_hours", "group_hours", "practical_hours",
                             "self_study_hours"]:
                    semester.setdefault("academic_hours", {})[key] = value
                elif key in ["term_papers", "modular_control_works", "calculation_graphic_works", "essays"]:
                    semester.setdefault("academic_task", {})[key] = value
                elif key in ["exam", "graded_test"] and value != 0:
                    semester.setdefault("reporting_type", REPORTING_TYPES[key])

        education_component["semesters"] = [semester for semester in education_component["semesters"] if
                                            all(semester.values())]

        table_data.append(education_component)
    return table_data


def processing_of_spreadsheet(path_to_file: str) -> list:
    """Processes spreadsheet file"""
    workbook = load_workbook(path_to_file, read_only=True, data_only=True)
    sheet_specialties = [specialty for specialty in workbook.sheetnames if re.match(r"\b\d+\b", specialty)]

    errors: list = []
    spreadsheet_result_data: list = []

    for sheet in sheet_specialties:
        worksheet = workbook[sheet]

        print_area = worksheet.print_area
        cell_range = str(print_area[0]).replace("$", "")
        start_cell, end_cell = get_start_and_end_cell_from_cell_range(cell_range)
        indexes, blocks = get_indexes_and_blocks_from_worksheet(worksheet, start_cell, end_cell)

        for index, block in zip(indexes, blocks):
            logger.trace(f"Block: {block}, index: {index}")
            header_block, table_block, footer_block = get_subblocks_from_block(worksheet, block, index)

            try:
                block_df = get_data_frame_from_cell_range(path_to_file, sheet, table_block)
                table_df = get_table_data_frame_from_block(block_df).dropna(axis=1, how="all")
                table_df = pd.DataFrame(table_df.values[1:], columns=table_df.iloc[0])
                table_df = table_df.loc[:, table_df.columns.notna()]

                header_data = get_data_from_header_block(worksheet[header_block])
                table_data = get_data_from_table_block(table_df)

                spreadsheet_result_data.append({
                    "specialty": header_data["Спеціальність"],
                    "specialization": header_data["Спеціалізація"],
                    "course_study": header_data["Курс навчання"],
                    "education_degree": header_data["Ступінь вищої освіти"],
                    "study_groups": header_data["Навчальна група"],
                    "education_components": table_data
                })
            except ParsingException as ex:
                ex.sheet_name = sheet
                ex.block = block
                logger.warning(ex)
                errors.append((ex.__str__(), sheet, block))
            except Exception as e:
                logger.error(e)

    workbook.close()

    filename = os.path.basename(path_to_file)
    logger.success(f"Spreadsheet '{filename}' parsed successfully")

    return spreadsheet_result_data, errors
